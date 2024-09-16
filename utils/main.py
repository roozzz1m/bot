import shutil
import os
import logging
from utils import DB
from playwright.async_api import async_playwright
from collections import Counter
from openpyxl import Workbook, load_workbook
import numpy as np
from datetime import datetime
from config import qw
import json
import psycopg2
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

title = PatternFill(start_color="00f18d", end_color="00f18d", fill_type="solid")
header_fill_1 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый
header_fill_2 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Синий
header_fill_3 = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Красный
header_fill_4 = PatternFill(start_color="bd99d6", end_color="bd99d6", fill_type="solid")
border_style = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
center_alignment = Alignment(horizontal="center", vertical="center")
font_bold = Font(bold=True)

def write_table(ws, start_col, start_row, data, header_fill):
    col_widths = {}
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            cell.border = border_style
            
            cell.alignment = center_alignment
            if row_idx == start_row:
                cell.font = font_bold
                cell.fill = header_fill
            
            col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(str(value)))
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = width + 2

async def get_links(page, url):
    await page.goto(url)
    await page.wait_for_selector('.audio_block_small_item__title a')
    links = await page.evaluate('''() => {
    return Array.from(document.querySelectorAll('.audio_block_small_item__title a')).map(link => {
        return {
            href: link.href,
            name: link.textContent.trim()
        };
    });
}
''')
    return links

async def mass(page, urls):
    generation = []
    for url in urls:
        url :str = url['href']
        url = url.replace("?ref=artist_related", '') + '/related'
        links = await get_links(page, url)
        generation.extend(links)

    top10 = []
    unique_generation = []
    seen = set()
    for item in generation:
        if (item['href'], item['name']) not in seen:
            unique_generation.append(item)
            seen.add((item['href'], item['name']))
            await DB().insert(table_name='users', values=(item['name'], item['href']))
        else:
            top10.append(item)

    return unique_generation, top10

async def get_generation(page, url):
    generation_1, top10_1 = await get_links(page, url), []
    print(generation_1)
    print('-'*79)
    generation_2, top10_2 = await mass(page, generation_1)
    print(generation_2)
    print('-'*79)
    generation_3, top10_3 = await mass(page, generation_2)
    print(generation_3)
    print('-'*79)

    top10 = [i for i in np.concatenate((top10_1, top10_2, top10_3), axis=None)]

    return generation_1, generation_2, generation_3, top10

async def get_name(url, page):
    await page.goto(url)
    print(url)
    name = await page.evaluate('''() => {
        const authorElement = document.querySelector('.MusicAuthor_block__title');
        if (authorElement) {
            return authorElement.textContent;
        }
        return 'Не найдено';
    }''')
    return name

async def main(url, save=True, id=None):
    logging.info(f"Начало работы с URL: {url}")
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            logging.info(f"Открываем страницу: {url}")

            name = await get_name(url, page)
            logging.info(f"Получено имя: {name}")

            url += '/related'
            gen1, gen2, gen3, top10 = await get_generation(page, url)

            # Сбор данных для записи в Excel
            data_1 = [['№', 'Имя артиста', 'Карточка артиста']]
            data_2 = [['№', 'Имя артиста', 'Карточка артиста']]
            data_3 = [['№', 'Имя артиста', 'Карточка артиста']]
            data_4 = [['№', 'Имя артиста', 'Карточка артиста']]
            data_title = [[f'{name} ({datetime.now().strftime("%d.%m.%Y")})']]

            c = 0
            for i in gen1:
                c += 1
                data_1.append([c, i['name'], i['href']])

            c = 0
            for i in gen2:
                c += 1
                data_2.append([c, i['name'], i['href']])

            c = 0
            for i in gen3:
                c += 1
                data_3.append([c, i['name'], i['href']])

            data = {}
            for i in top10:
                data[i['href']] = i['name']

            counter = Counter([i['href'] for i in top10])

            top_10 = counter.most_common(10)

            c = 0
            for i in top_10:
                c += 1
                data_4.append([c, data[i[0]], i[0]])

            # Сохранение данных в Excel
            file_name = "last.xlsx"
            if os.path.exists(file_name) and save:
                wb = load_workbook(file_name)
                ws = wb.create_sheet(title=f'{name} ({datetime.now().strftime("%d.%m.%Y")})')
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = f'{name} ({datetime.now().strftime("%d.%m.%Y")})'

            write_table(ws, 1, 1, data_title, title)
            write_table(ws, 2, 2, data_1, header_fill_1)
            write_table(ws, 6, 2, data_2, header_fill_2)
            write_table(ws, 10, 2, data_3, header_fill_3)
            write_table(ws, 14, 2, data_4, header_fill_4)

            filename = f"{name}-{datetime.now().strftime('%d%m%Y')}.xlsx"
            wb.save(filename)
            logging.info(f"Файл сохранен: {filename}")

            if not save:
                source = "artists.xlsx"
                destination = 'last.xlsx'
                shutil.copy(source, destination)

            # Обновление даты в settings.json
            with open('settings.json', 'r') as f:
                data = json.load(f)
                data['date'] = datetime.now().strftime("%d.%m.%Y")

            with open('settings.json', 'w') as f:
                json.dump(data, f, indent=4)

            # qw.remove(id)
            await browser.close()
            logging.info("Задача успешно завершена")

            return filename

    except Exception as e:
        logging.error(f"Ошибка во время выполнения main: {e}")
        raise
