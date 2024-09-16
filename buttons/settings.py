import utils
from aiogram.types import Message
from aiogram.fsm.context import FSMContext
from aiogram.types import WebAppInfo
from aiogram.filters import CommandObject, Command
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.enums import ChatMemberStatus
import json
from openpyxl import load_workbook
from config import qw

from misc import dp, bot
from aiogram import F, types

@dp.callback_query(F.data == "settings")
async def settings(callback: types.CallbackQuery, state: FSMContext):

    with open('settings.json', 'r', encoding='utf-8') as file:
        data :dict = json.load(file)

    text_btn = "❌ Сохранять файл"
    if data['save'] is True:
        text_btn = "✅ Сохранять файл"

    btn = InlineKeyboardBuilder().row(types.InlineKeyboardButton(text=text_btn, callback_data='save'), types.InlineKeyboardButton(text="🗑️ Очистить файл", callback_data='clear')).row(types.InlineKeyboardButton(text='⭐ Очистить очередь', callback_data='clear_qw')).row(types.InlineKeyboardButton(text='⭐ В главное меню', callback_data='menu')).as_markup()
    await callback.message.edit_text(text='Укажи своё действие:', reply_markup=btn)

@dp.callback_query(F.data == "clear_qw")
async def clear_qw(callback: types.CallbackQuery, state: FSMContext):
    qw.clear()
    await callback.answer("✅",show_alert=True)


@dp.callback_query(F.data == "menu")
async def menu(callback: types.CallbackQuery, state: FSMContext):
    with open('settings.json', 'r') as f:
        data = json.load(f)

    text = f"Добро пожаловать, укажи своё действие.\nПоследние обновление данных в файле было: {data['date']}"
    btn = InlineKeyboardBuilder().add(types.InlineKeyboardButton(text='🚀 Начать', callback_data='start')).row(types.InlineKeyboardButton(text='🪬 Получить последний файл', callback_data="last_file"), types.InlineKeyboardButton(text="⚙️ Настройки", callback_data='settings')).as_markup()
    await state.set_state(utils.Start.url)
    await callback.message.edit_text(text=text, reply_markup=btn)

def clear_(file_path):
    workbook = load_workbook(file_path)

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.delete_cols(1, worksheet.max_column)

    workbook.save(file_path)

@dp.callback_query(F.data == "clear")
async def clear(callback: types.CallbackQuery, state: FSMContext):
    for i in ['artists.xlsx', 'last.xlsx']:
        try:
            clear_(i)
        except:
            pass

    await callback.answer("✅",show_alert=True)

@dp.callback_query(F.data == "save")
async def save(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer(text='Команда выключена', show_alert=True)
#     with open('settings.json', 'r') as f:
#         data = json.load(f)
        
#         if data['save'] is True:
#             data['save'] = False
#         else:
#             data['save'] = True
        
#         with open('settings.json', 'w') as f:
#             json.dump(data, f, indent=4)

#     text_btn = "❌ Сохранять файл"
#     if data['save'] is True:
#         text_btn = "✅ Сохранять файл"

#     btn = InlineKeyboardBuilder().row(types.InlineKeyboardButton(text=text_btn, callback_data='save'), types.InlineKeyboardButton(text="🗑️ Очистить файл", callback_data='clear')).row(types.InlineKeyboardButton(text='⭐ Очистить очередь', callback_data='clear_qw')).row(types.InlineKeyboardButton(text='⭐ В главное меню', callback_data='menu')).as_markup()
#     await callback.message.edit_reply_markup(reply_markup=btn)
